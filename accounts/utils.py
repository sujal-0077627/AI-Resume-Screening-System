import pdfplumber
import spacy
import re
import csv
import pandas as pd
from django.http import HttpResponse
from django.conf import settings
from django.core.mail import send_mail
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

nlp = spacy.load("en_core_web_sm")

SKILLS_DATABASE = [
    'python', 'java', 'javascript', 'typescript', 'c', 'c++', 'c#', 'ruby',
    'php', 'swift', 'kotlin', 'go', 'rust', 'scala', 'r', 'matlab', 'perl',
    'html', 'css', 'html5', 'css3', 'react', 'angular', 'vue', 'node',
    'django', 'flask', 'fastapi', 'spring', 'express', 'laravel',
    'bootstrap', 'tailwind', 'sass', 'less', 'redux', 'webpack', 'babel',
    'sql', 'mysql', 'postgresql', 'mongodb', 'sqlite', 'oracle', 'redis',
    'cassandra', 'dynamodb', 'elasticsearch', 'mariadb', 'firebase',
    'pandas', 'numpy', 'scipy', 'matplotlib', 'seaborn', 'plotly',
    'scikit-learn', 'tensorflow', 'pytorch', 'keras', 'opencv',
    'nltk', 'spacy', 'transformers', 'huggingface', 'openai',
    'machine learning', 'deep learning', 'data science', 'data analysis',
    'data visualization', 'natural language processing', 'nlp',
    'computer vision', 'artificial intelligence', 'generative ai',
    'reinforcement learning', 'predictive modeling', 'statistics',
    'git', 'github', 'gitlab', 'bitbucket', 'docker', 'kubernetes',
    'jenkins', 'aws', 'azure', 'gcp', 'linux', 'ubuntu', 'bash',
    'shell', 'powershell', 'ansible', 'terraform', 'jira', 'agile',
    'scrum', 'rest api', 'graphql', 'soap', 'microservices',
    'ci/cd', 'devops', 'testing', 'selenium', 'pytest', 'unittest',
    'junit', 'postman', 'tableau', 'power bi', 'excel', 'hadoop',
    'spark', 'kafka', 'airflow', 'etl', 'data warehousing', 'big data',
    'celery', 'rabbitmq', 'gunicorn', 'nginx', 'apache',
    'django rest framework', 'websocket', 'tornado',
]


def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print(f"Error reading PDF: {e}")
    return text


def extract_email(text):
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    match = re.search(email_pattern, text)
    return match.group(0) if match else ""


def extract_phone(text):
    phone_pattern = r'(\+?\d{10,13}|\+?\d{2,4}[\s\-]?\d{6,8})'
    match = re.search(phone_pattern, text)
    return match.group(0).strip() if match else ""


def extract_experience(text):
    years = 0
    patterns = [
        r'(\d+)\+?\s*years?\s*(?:of\s*)?experience',
        r'experience\s*(?:of\s*)?(\d+)\+?\s*years?',
        r'(\d+)\+?\s*yrs?\s*(?:of\s*)?experience',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            years = int(match.group(1))
            break
    
    if re.search(r'fresher', text, re.IGNORECASE):
        years = 0
    
    if years <= 1:
        level = 'Fresher'
    elif years <= 3:
        level = 'Junior'
    else:
        level = 'Senior'
    
    return years, level


def extract_name_nlp(text):
    """Find the candidate's name from resume text.
    Step 1: Look for an explicit 'Name: XYZ' style label (very reliable when present).
    Step 2: Fall back to spaCy NER (PERSON entity) near the top of the resume.
    """
    if not text:
        return None

    top_section = text.strip()[:600]

    # Step 1 — explicit "Name :" / "Name -" / "Full Name:" label (common in resumes)
    label_match = re.search(
        r'(?:full\s*name|name)\s*[:\-]\s*([A-Za-z][A-Za-z.\'\-]*(?:[ \t]+[A-Za-z][A-Za-z.\'\-]*){0,3})',
        top_section, re.IGNORECASE
    )
    if label_match:
        candidate = label_match.group(1).strip()
        words = candidate.split()
        if 1 < len(words) <= 4:
            return candidate.title()

    # Step 2 — spaCy NER fallback (works well for resumes without an explicit "Name:" label)
    doc = nlp(top_section)
    for ent in doc.ents:
        if ent.label_ == "PERSON":
            name = ent.text.strip()
            if 1 < len(name.split()) <= 4:
                return name.title()

    return None


def extract_skills(text):
    text_lower = text.lower()
    found_skills = set()

    for skill in SKILLS_DATABASE:
        pattern = r'(?<![a-z0-9])' + re.escape(skill) + r'(?![a-z0-9])'
        if re.search(pattern, text_lower):
            found_skills.add(skill)

    return sorted(found_skills)


_semantic_model = None
_semantic_model_failed = False


def _get_semantic_model():
    """Lazily loads the Sentence-Transformers model once. Never crashes the app if
    the package/model isn't available (e.g. no internet) — just disables the semantic boost."""
    global _semantic_model, _semantic_model_failed
    if _semantic_model_failed:
        return None
    if _semantic_model is None:
        try:
            from sentence_transformers import SentenceTransformer
            print("[Sentence-Transformers] Loading model 'all-MiniLM-L6-v2' (first time only)...")
            _semantic_model = SentenceTransformer('all-MiniLM-L6-v2')
            print("[Sentence-Transformers] Model loaded.")
        except Exception as e:
            print(f"[Sentence-Transformers] Could not load model, semantic matching disabled: {e}")
            _semantic_model_failed = True
            return None
    return _semantic_model


def get_semantic_similarity(resume_text, job_description):
    """Returns a 0-100 semantic similarity score between resume and job description,
    or None if Sentence-Transformers isn't available (caller should fall back gracefully)."""
    model = _get_semantic_model()
    if not model:
        return None
    try:
        from sentence_transformers import util
        embeddings = model.encode(
            [resume_text[:2000], job_description[:2000]],
            convert_to_tensor=True
        )
        similarity = util.cos_sim(embeddings[0], embeddings[1]).item()
        return round(max(0.0, similarity) * 100, 2)
    except Exception as e:
        print(f"[Sentence-Transformers] Similarity calculation failed: {e}")
        return None


def get_tfidf_similarity(resume_text, job_description):
    """Real TF-IDF + Cosine Similarity between resume and job description.
    Converts both texts into TF-IDF vectors (word importance scores) and measures
    how similar those vectors are (0-100). Returns None only if something goes wrong,
    so the caller can fall back safely."""
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity

        documents = [resume_text, job_description]
        vectorizer = TfidfVectorizer(stop_words='english')
        tfidf_matrix = vectorizer.fit_transform(documents)

        similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
        return round(max(0.0, similarity) * 100, 2)
    except Exception as e:
        print(f"[TF-IDF] Similarity calculation failed: {e}")
        return None


def calculate_match_score(resume_text, job_description):
    if not resume_text or not job_description:
        return 0.0, [], []

    job_skills = extract_skills(job_description)
    resume_skills = extract_skills(resume_text)

    if not job_skills:
        return 0.0, [], []

    job_set = set(job_skills)
    resume_set = set(resume_skills)

    matched = sorted(job_set.intersection(resume_set))
    missing = sorted(job_set - resume_set)

    keyword_score = (len(matched) / len(job_set)) * 100
    tfidf_score = get_tfidf_similarity(resume_text, job_description)
    semantic_score = get_semantic_similarity(resume_text, job_description)  # None unless Sentence-Transformers is installed

    if semantic_score is not None:
        # All three available: skill match + TF-IDF + semantic
        score = round((keyword_score * 0.4) + (tfidf_score * 0.3) + (semantic_score * 0.3), 2)
    elif tfidf_score is not None:
        # Skill match (explainable) + TF-IDF + Cosine Similarity (the required technique)
        score = round((keyword_score * 0.5) + (tfidf_score * 0.5), 2)
    else:
        # Safe fallback if scikit-learn isn't available for some reason
        score = round(keyword_score, 2)

    return score, matched, missing


def get_candidate_full_details(candidate):
    """Candidate aur Score model se complete data nikalne ke liye helper"""
    score_obj = candidate.score_set.first()
    score_val = score_obj.matching_score if score_obj else 0.0

    if score_val >= 70:
        calculated_status = "Shortlisted"
    elif score_val >= 40:
        calculated_status = "Under Review"
    else:
        calculated_status = "Not Recommended"

    # Merge Experience (Years + Level) into 1 clean string
    exp_years = getattr(candidate, 'experience_years', 0)
    exp_level = getattr(candidate, 'experience_level', 'Fresher')
    formatted_exp = f"{exp_level} ({exp_years} Yrs)" if exp_years > 0 else f"{exp_level}"

    return {
        'Candidate Name': getattr(candidate, 'name', 'N/A'),
        'Email': getattr(candidate, 'email', 'N/A'),
        'Phone': getattr(candidate, 'phone', 'N/A'),
        'Applied Job': candidate.applied_job.title if candidate.applied_job else 'N/A',
        'Experience': formatted_exp,
        'Match Score': f"{score_val}%",
        'Status': calculated_status,
        'Matched Skills': getattr(score_obj, 'matched_skills', 'N/A') if score_obj else 'N/A',
        'Missing Skills': getattr(score_obj, 'missing_skills', 'N/A') if score_obj else 'N/A',
        'AI Explanation': getattr(score_obj, 'explanation', 'N/A') if score_obj else 'N/A',
        'Applied Date': candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else 'N/A',
    }


def export_candidates_csv(queryset):
    """CSV Export"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="candidates_full_list.csv"'
    
    writer = csv.writer(response)
    
    if queryset.exists():
        first_candidate_data = get_candidate_full_details(queryset.first())
        writer.writerow(first_candidate_data.keys())
        
        for candidate in queryset:
            data = get_candidate_full_details(candidate)
            writer.writerow(data.values())
            
    return response


def export_candidates_excel(queryset):
    """Excel Export with Professional Formatting (Auto Width, Colors, Borders)"""
    data_list = [get_candidate_full_details(c) for c in queryset]
    df = pd.DataFrame(data_list)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidates_full_list.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Candidates Details')
        
        # Openpyxl worksheet styling
        worksheet = writer.sheets['Candidates Details']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF") # White Bold Text
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Style Header Row
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit Column Widths & Set Borders
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            # Cap width between 14 and 45 for clean reading
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)
            
            for cell in col:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="center")

    return response


# =========================================================
# GENAI SHORTLIST / REJECT EMAIL
# =========================================================
SHORTLIST_THRESHOLD = 80  # score >= this = shortlisted


def _generate_email_body(candidate_name, job_title, status, score, matched_skills, missing_skills):
    """Use Gemini to write the email body. Falls back to a fixed template if the API fails."""
    try:
        import google.generativeai as genai
        genai.configure(api_key=settings.GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-flash-latest')

        if status == 'Shortlisted':
            prompt = (
                f"Write a warm, professional email to a job candidate named {candidate_name} "
                f"informing them they have been SHORTLISTED for the '{job_title}' position. "
                f"Their resume match score was {score}%. Their matching skills: {matched_skills or 'relevant skills'}. "
                f"Mention 2-3 of these matched skills naturally, congratulate them, and say the hiring team "
                f"will reach out soon with next steps. Keep it under 130 words. "
                f"Output ONLY the email body text, no subject line, no placeholders like [Your Name]."
            )
        else:
            prompt = (
                f"Write a polite, respectful, encouraging rejection email to a job candidate named {candidate_name} "
                f"for the '{job_title}' position. Their resume match score was {score}%. "
                f"The specific skills they were missing: {missing_skills or 'some required skills'}. "
                f"Mention the missing skills as the constructive reason, encourage them to develop these skills "
                f"and apply again in the future. Keep it under 130 words, positive tone. "
                f"Output ONLY the email body text, no subject line, no placeholders like [Your Name]."
            )

        response = model.generate_content(prompt, request_options={'timeout': 15})
        text = response.text.strip() if response and response.text else ""
        if text:
            return text
    except Exception as e:
        print(f"[GenAI Email] Gemini API failed, using fallback template: {e}")

    # Fallback (used if Gemini API fails / no internet / quota over)
    if status == 'Shortlisted':
        return (
            f"Dear {candidate_name},\n\n"
            f"Congratulations! You have been shortlisted for the {job_title} position "
            f"with a match score of {score}%. Your skills in {matched_skills or 'the required areas'} "
            f"align well with what we are looking for.\n\n"
            f"Our team will reach out soon with the next steps.\n\n"
            f"Best regards,\nHiring Team"
        )
    else:
        return (
            f"Dear {candidate_name},\n\n"
            f"Thank you for applying for the {job_title} position. After careful review, "
            f"we found a gap in the following skills required for this role: {missing_skills or 'a few key areas'}.\n\n"
            f"We encourage you to build on these skills and apply again in the future.\n\n"
            f"Best regards,\nHiring Team"
        )


def send_candidate_status_email(candidate, job, score_val, matched_skills, missing_skills):
    """Automatically emails the candidate Shortlisted/Rejected based on SHORTLIST_THRESHOLD."""
    if not candidate.email or candidate.email.startswith('noemail_'):
        print(f"[GenAI Email] Skipped {candidate.name} — no valid email extracted from resume.")
        return

    status = 'Shortlisted' if score_val >= SHORTLIST_THRESHOLD else 'Rejected'
    subject = f"Application Update: {job.title} — {status}"
    body = _generate_email_body(candidate.name, job.title, status, score_val, matched_skills, missing_skills)

    try:
        send_mail(
            subject,
            body,
            settings.DEFAULT_FROM_EMAIL,
            [candidate.email],
            fail_silently=False,
        )
        print(f"[GenAI Email] {status} email sent to {candidate.email}")
    except Exception as e:
        print(f"[GenAI Email] Failed to send email to {candidate.email}: {e}")